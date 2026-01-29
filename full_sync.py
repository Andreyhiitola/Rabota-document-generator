#!/usr/bin/env python3
"""
full_sync.py - Полная синхронизация Trello ↔ Excel ↔ Dropbox
С АВТОСОРТИРОВКОЙ по дате начала работ
"""
import os
import subprocess
import shutil
import dropbox
import requests
import openpyxl
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

def get_dropbox_client():
    """Получает access токен из refresh токена"""
    print("🔄 Dropbox: refresh токен → access токен...")
    
    app_key = os.getenv('DROPBOX_APP_KEY')
    app_secret = os.getenv('DROPBOX_APP_SECRET')
    refresh_token = os.getenv('DROPBOX_REFRESH_TOKEN')
    
    if not all([app_key, app_secret, refresh_token]):
        print("❌ DROPBOX_* переменные не найдены в .env")
        exit(1)
    
    response = requests.post('https://api.dropbox.com/oauth2/token', data={
        'grant_type': 'refresh_token',
        'refresh_token': refresh_token,
        'client_id': app_key,
        'client_secret': app_secret,
    })
    
    if response.status_code != 200:
        print(f"❌ Dropbox API: {response.status_code}")
        print(response.text)
        exit(1)
    
    data = response.json()
    print(f"✅ Access токен получен")
    return dropbox.Dropbox(data['access_token'])

def sort_excel_by_date(excel_file: str):
    """
    Сортировка Excel по дате начала работ (колонка D)
    Старые работы сверху, новые внизу
    """
    print("\n📊 ШАГ 5/5: Сортировка строк по дате")
    print("-" * 80)
    
    try:
        from copy import copy
        from dateutil import parser as date_parser
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        if ws.max_row < 3:
            print("  ℹ️  Нечего сортировать (мало строк)")
            return
        
        print(f"  📋 Сортировка {ws.max_row - 1} строк данных...")
        
        # Функция для нормализации даты
        def normalize_date(date_val):
            """Преобразует любое значение даты в datetime или None"""
            if date_val is None:
                return None
            if isinstance(date_val, datetime):
                return date_val
            if isinstance(date_val, str):
                try:
                    return date_parser.parse(date_val, dayfirst=True)
                except:
                    return None
            return None
        
        # Собираем все строки с данными
        rows_data = []
        for row_num in range(2, ws.max_row + 1):
            # Читаем всю строку
            row_values = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row_num, col)
                row_values.append({
                    'value': cell.value,
                    'font': copy(cell.font) if cell.font else None,
                    'fill': copy(cell.fill) if cell.fill else None,
                    'border': copy(cell.border) if cell.border else None,
                    'alignment': copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format,
                })
            
            # Получаем и нормализуем дату начала работ (колонка 4 = D)
            start_date_raw = ws.cell(row_num, 4).value
            start_date = normalize_date(start_date_raw)
            
            rows_data.append({
                'start_date': start_date,
                'values': row_values
            })
        
        # Сортируем: сначала по дате (старые сверху), строки без даты в конец
        rows_data.sort(key=lambda r: (
            r['start_date'] is None,  # None в конец
            r['start_date'] if r['start_date'] is not None else datetime.max
        ))
        
        # Записываем отсортированные строки обратно
        for new_row_num, row_data in enumerate(rows_data, start=2):
            for col_num, cell_data in enumerate(row_data['values'], start=1):
                cell = ws.cell(new_row_num, col_num)
                cell.value = cell_data['value']
                
                if cell_data['font']:
                    cell.font = cell_data['font']
                if cell_data['fill']:
                    cell.fill = cell_data['fill']
                if cell_data['border']:
                    cell.border = cell_data['border']
                if cell_data['alignment']:
                    cell.alignment = cell_data['alignment']
                if cell_data['number_format']:
                    cell.number_format = cell_data['number_format']
        
        # Сохраняем
        wb.save(excel_file)
        
        print(f"  ✅ Строки отсортированы!")
        print(f"     📅 Старые работы → сверху")
        print(f"     📅 Новые работы → внизу")
        
        # Показываем первые 3 даты для проверки
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        print(f"\n  🔍 Первые 3 даты начала работ:")
        for row in range(2, min(5, ws.max_row + 1)):
            date_val = ws.cell(row, 4).value
            address = ws.cell(row, 3).value
            if address:
                address_short = str(address)[:50] + "..." if len(str(address)) > 50 else str(address)
            else:
                address_short = "Нет адреса"
            print(f"     Строка {row}: {date_val} - {address_short}")
        
    except Exception as e:
        print(f"  ⚠️ Ошибка сортировки: {e}")
        print(f"  Файл сохранён без сортировки")

def full_sync():
    print("="*80)
    print(f"🚀 ПОЛНАЯ СИНХРОНИЗАЦИЯ TRELLO → EXCEL → DROPBOX + СОРТИРОВКА")
    print(f"⏰ Начало: {datetime.now()}")
    print("="*80)
    
    # Подключение к Dropbox
    dbx = get_dropbox_client()
    
    # Пути к файлам
    dropbox_path = "/data.xlsx"  # ✅ Файл в КОРНЕ Dropbox (40 KB)
    base_excel = "data.xlsx"      # Локальная копия в корне проекта
    tmp_excel = "/tmp/data.xlsx"  # Временный файл для обработки
    
    # Путь к sync_trello_severen.py
    sync_script = "severen-generator/sync_trello_severen.py"
    if not os.path.exists(sync_script):
        sync_script = "sync_trello_severen.py"
        if not os.path.exists(sync_script):
            print(f"❌ sync_trello_severen.py не найден!")
            exit(1)
    
    # === ШАГ 1: Скачиваем актуальный файл из Dropbox ===
    print("\n📥 ШАГ 1/5: Скачивание data.xlsx из Dropbox (корень)")
    print("-" * 80)
    
    try:
        metadata, response = dbx.files_download(dropbox_path)
        with open(base_excel, 'wb') as f:
            f.write(response.content)
        
        size = os.path.getsize(base_excel)
        print(f"  ✅ Скачан: {dropbox_path} → {base_excel}")
        print(f"     📊 Размер: {size / 1024:.2f} KB")
        print(f"     ⏰ Изменён в Dropbox: {metadata.server_modified}")
        
        if size < 1024:
            print(f"  ⚠️ ВНИМАНИЕ! Файл подозрительно маленький ({size} байт)!")
            
    except dropbox.exceptions.ApiError as e:
        print(f"  ❌ Файл {dropbox_path} не найден в Dropbox: {e}")
        exit(1)
    
    # === ШАГ 2: Копируем в /tmp для обработки ===
    print("\n📋 ШАГ 2/5: Подготовка файла для Trello")
    print("-" * 80)
    
    shutil.copy(base_excel, tmp_excel)
    stat = os.stat(tmp_excel)
    print(f"  ✅ Скопирован: {base_excel} → {tmp_excel}")
    print(f"     📊 Размер: {stat.st_size / 1024:.2f} KB")
    
    # === ШАГ 3: Trello синхронизация ===
    print("\n🔄 ШАГ 3/5: Синхронизация с Trello")
    print("-" * 80)
    print(f"  📝 Используется скрипт: {sync_script}")
    
    stat_before = os.stat(tmp_excel)
    print(f"  📄 ДО sync_trello: {stat_before.st_size / 1024:.2f} KB")
    
    try:
        result = subprocess.run(
            ['python3', sync_script, '--file', tmp_excel],
            check=True, capture_output=True, text=True
        )
        
        # Показываем вывод скрипта
        if result.stdout.strip():
            for line in result.stdout.strip().split('\n'):
                print(f"    {line}")
        
        stat_after = os.stat(tmp_excel)
        print(f"  📄 ПОСЛЕ sync_trello: {stat_after.st_size / 1024:.2f} KB")
        
        size_diff = stat_after.st_size - stat_before.st_size
        if size_diff != 0:
            print(f"  ✅ Файл изменён! Δ = {size_diff / 1024:+.2f} KB")
        else:
            print(f"  ℹ️  Размер не изменился (нет новых данных из Trello)")
        
    except subprocess.CalledProcessError as e:
        print(f"  ❌ Ошибка sync_trello_severen.py:")
        if e.stdout:
            print(e.stdout)
        if e.stderr:
            print(e.stderr)
        exit(1)
    
    # === ШАГ 4: Сортировка по дате ===
    sort_excel_by_date(tmp_excel)
    
    # === ШАГ 5: Загрузка обратно в Dropbox ===
    print("\n📤 ШАГ 6/6: Загрузка обновлённого файла в Dropbox")
    print("-" * 80)
    
    stat = os.stat(tmp_excel)
    print(f"  📄 Обновлённый файл: {tmp_excel}")
    print(f"     📊 Размер: {stat.st_size / 1024:.2f} KB")
    print(f"     ⏰ Изменён: {datetime.fromtimestamp(stat.st_mtime)}")
    
    with open(tmp_excel, 'rb') as f:
        dbx.files_upload(
            f.read(), 
            dropbox_path,
            mode=dropbox.files.WriteMode('overwrite')
        )
    
    print(f"  ✅ Загружен в Dropbox: {dropbox_path}")
    
    # Обновляем локальную копию
    shutil.copy(tmp_excel, base_excel)
    print(f"  ✅ Обновлён локально: {base_excel}")
    
    print("\n" + "="*80)
    print(f"✅ УСПЕХ! Синхронизация и сортировка завершены")
    print(f"⏰ Завершено: {datetime.now()}")
    print("="*80)

if __name__ == '__main__':
    try:
        full_sync()
    except KeyboardInterrupt:
        print("\n⏹️ Остановлено пользователем")
    except Exception as e:
        print(f"\n❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
