#!/usr/bin/env python3
"""
Скрипт синхронизации данных из Trello в Excel файл Rabochie-tabl.-SMR-v2.xlsx
Парсит описания карточек Trello и записывает в правильные колонки
"""

import os
import re
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

TRELLO_API_KEY = os.getenv('TRELLO_API_KEY')
TRELLO_TOKEN = os.getenv('TRELLO_TOKEN')
TRELLO_BOARD_ID = os.getenv('TRELLO_BOARD_ID')
EXCEL_FILE = '/app/files/Рабочие_табл_СМР_v2.xlsx'  # Путь в Docker

# Колонки в Excel (индексы 1-15)
COLUMNS = {
    'hide': 1,           # A: Скрыть (x)
    'act_type': 2,       # B: Акты на закрытие 1
    'act_status': 3,     # C: Акты на закрытие 2
    'address': 4,        # D: Адрес + Задание
    'start_date': 5,     # E: Начало работ
    'end_date': 6,       # F: Конец работ
    'client': 7,         # G: Клиент
    'contractor': 8,     # H: Исполнитель (ПО)
    'status': 9,         # I: Статус
    'work_name': 10,     # J: Название работ
    'cost': 11,          # K: Стоимость (руб) - автоматически
    'report_date': 12,   # L: Дата формирования отчета
    'transit': 13,       # M: Транзитные адреса
    'note': 14,          # N: Примечание
    'trello_desc': 15    # O: Описание из Trello (архив)
}


def parse_card_description(description):
    """
    Парсит описание карточки Trello и извлекает структурированные данные
    """
    data = {
        'district': '',
        'address': '',
        'start_date': '',
        'client': '',
        'contractor': '',
        'note': '',
        'full_description': description
    }

    # Парсим район (первая строка с номером)
    district_match = re.search(r'^\d+\n(.+?)\n', description, re.MULTILINE)
    if district_match:
        data['district'] = district_match.group(1).strip()

    # Парсим дату начала работ
    start_date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+Начало работ', description, re.IGNORECASE)
    if start_date_match:
        date_str = start_date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['start_date'] = date_obj.strftime('%Y-%m-%d')
        except:
            data['start_date'] = date_str

    # Парсим клиента
    client_match = re.search(r'Клиент\s+\[?([^\]\n]+?)\]?(?:\(|,|\n)', description, re.IGNORECASE)
    if client_match:
        data['client'] = client_match.group(1).strip()

    # Парсим заказчика (альтернативное поле)
    if not data['client']:
        client_match = re.search(r'Заказчик:\s+\[?([^\]\n]+?)\]?(?:\(|,|\n)', description, re.IGNORECASE)
        if client_match:
            data['client'] = client_match.group(1).strip()

    # Парсим подрядчика
    contractor_match = re.search(r'Подрядчик:\s+(.+?)(?:\n|$)', description, re.IGNORECASE)
    if contractor_match:
        contractor = contractor_match.group(1).strip()
        if contractor and contractor != '?????':
            data['contractor'] = contractor

    # Парсим адрес (гор. Санкт-Петербург...)
    address_match = re.search(r'гор\.?\s+Санкт-Петербург[^\n]+', description, re.IGNORECASE)
    if address_match:
        data['address'] = address_match.group(0).strip()

    # Если нет полного адреса, пробуем просто адрес
    if not data['address']:
        address_match = re.search(r'(?:по адресу[:\s]+|адрес[:\s]+)([^\n]+)', description, re.IGNORECASE)
        if address_match:
            addr = address_match.group(1).strip()
            if data['district']:
                data['address'] = f"гор. Санкт-Петербург, {data['district']}, {addr}"
            else:
                data['address'] = f"гор. Санкт-Петербург, {addr}"

    return data


def get_trello_cards():
    """
    Получает все карточки с доски Trello
    """
    url = f"https://api.trello.com/1/boards/{TRELLO_BOARD_ID}/cards"
    params = {
        'key': TRELLO_API_KEY,
        'token': TRELLO_TOKEN,
        'fields': 'id,name,desc,idList,labels,due'
    }

    response = requests.get(url, params=params)
    response.raise_for_status()
    return response.json()


def get_list_name(list_id):
    """
    Получает название листа Trello по ID
    """
    url = f"https://api.trello.com/1/lists/{list_id}"
    params = {
        'key': TRELLO_API_KEY,
        'token': TRELLO_TOKEN,
        'fields': 'name'
    }

    response = requests.get(url, params=params)
    response.raise_for_status()
    return response.json()['name']


def determine_status(list_name):
    """
    Определяет статус работы по названию листа Trello
    """
    list_name_lower = list_name.lower()

    if 'выполнен' in list_name_lower or 'done' in list_name_lower:
        return 'Выполнен'
    elif 'приостановк' in list_name_lower or 'pause' in list_name_lower:
        return 'Приостановка'
    elif 'отказ' in list_name_lower or 'cancel' in list_name_lower:
        return 'Отказ'
    else:
        return 'В работе'


def sync_to_excel():
    """
    Синхронизирует данные из Trello в Excel
    """
    print("🔄 Начинаю синхронизацию Trello → Excel")

    # Проверяем наличие файла
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Файл не найден: {EXCEL_FILE}")
        return

    # Получаем карточки из Trello
    print("📥 Получаю карточки из Trello...")
    cards = get_trello_cards()
    print(f"✅ Найдено карточек: {len(cards)}")

    # Загружаем Excel
    print(f"📂 Загружаю файл {EXCEL_FILE}...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Работы']

    # Находим следующую пустую строку
    next_row = ws.max_row + 1

    # Обрабатываем каждую карточку
    added_count = 0
    for card in cards:
        if not card.get('desc'):
            continue

        # Парсим описание
        data = parse_card_description(card['desc'])

        # Пропускаем карточки без адреса
        if not data['address']:
            continue

        # Получаем статус из листа Trello
        list_name = get_list_name(card['idList'])
        status = determine_status(list_name)

        # Записываем в Excel
        ws.cell(next_row, COLUMNS['address'], data['address'])
        ws.cell(next_row, COLUMNS['start_date'], data['start_date'])
        ws.cell(next_row, COLUMNS['client'], data['client'])
        ws.cell(next_row, COLUMNS['contractor'], data['contractor'])
        ws.cell(next_row, COLUMNS['status'], status)
        ws.cell(next_row, COLUMNS['note'], data['note'])
        ws.cell(next_row, COLUMNS['trello_desc'], data['full_description'])

        print(f"✅ Добавлена строка {next_row}: {data['address'][:50]}...")

        next_row += 1
        added_count += 1

    # Сохраняем файл
    print(f"💾 Сохраняю изменения...")
    wb.save(EXCEL_FILE)
    print(f"✅ Синхронизация завершена! Добавлено строк: {added_count}")


if __name__ == '__main__':
    try:
        sync_to_excel()
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
