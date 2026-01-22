#!/usr/bin/env python3
"""
ГЕНЕРАТОР ОТЧЕТОВ ДЛЯ НОВОЙ СТРУКТУРЫ РАБОЧЕЙ ТАБЛИЦЫ
Версия: 2.0
Дата: 22.01.2026

Новая структура таблицы:
- A: Скрыть (x)
- B: Акты на закрытие 1
- C: Акты на закрытие 2
- D: Адрес + Задание (полный адрес)
- E: Начало работ
- F: Конец работ
- G: Клиент
- H: Исполнитель (ПО)
- I: Статус
- J: Название работ
- K: Стоимость (руб)
- L: Дата формирования отчета
- M: Транзитные адреса
- N: Примечание
- O: Описание из Trello (архив)
"""

import os
import sys
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Ошибка: openpyxl не установлен")
    print("Установите: pip install openpyxl")
    sys.exit(1)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ReportGeneratorV2:
    """Генератор отчетов для новой структуры таблицы"""

    def __init__(self, source_file: str):
        self.source_file = source_file
        self.wb = None
        self.source_ws = None

        # НОВЫЕ ИНДЕКСЫ КОЛОНОК (1-based)
        self.COL_HIDE = 1      # A - Скрыть (x)
        self.COL_ACT1 = 2      # B - Акты 1
        self.COL_ACT2 = 3      # C - Акты 2
        self.COL_ADDR = 4      # D - Адрес + Задание (ПОЛНЫЙ!)
        self.COL_START = 5     # E - Начало работ
        self.COL_END = 6       # F - Конец работ
        self.COL_CLIENT = 7    # G - Клиент
        self.COL_EXEC = 8      # H - Исполнитель (ПО)
        self.COL_STATUS = 9    # I - Статус
        self.COL_WORK = 10     # J - Название работ
        self.COL_PRICE = 11    # K - Стоимость
        self.COL_DATE = 12     # L - Дата формирования отчета
        self.COL_TRANS = 13    # M - Транзитные адреса
        self.COL_NOTE = 14     # N - Примечание
        self.COL_DESC = 15     # O - Описание Trello

    def load_source(self) -> bool:
        """Загрузить исходную таблицу"""
        logger.info(f"📂 Загрузка файла: {self.source_file}")

        if not os.path.exists(self.source_file):
            logger.error(f"❌ Файл не найден: {self.source_file}")
            return False

        try:
            self.wb = openpyxl.load_workbook(self.source_file)

            # Ищем лист "Работы"
            if "Работы" in self.wb.sheetnames:
                self.source_ws = self.wb["Работы"]
            elif "Data" in self.wb.sheetnames:
                self.source_ws = self.wb["Data"]
            else:
                self.source_ws = self.wb.active

            logger.info(f"✅ Загружен лист: {self.source_ws.title}")
            return True

        except Exception as e:
            logger.error(f"❌ Ошибка загрузки: {e}")
            return False

    def extract_data(
        self, 
        start_date: str = None, 
        end_date: str = None,
        filter_status: str = None,
        filter_executor: str = None
    ) -> List[Dict]:
        """
        Извлечь данные из таблицы с фильтрацией

        Args:
            start_date: Дата начала периода (YYYY-MM-DD)
            end_date: Дата окончания периода (YYYY-MM-DD)
            filter_status: Фильтр по статусу (В работе, Выполнен, и т.д.)
            filter_executor: Фильтр по исполнителю
        """
        logger.info("🔍 Извлечение данных из таблицы...")
        data = []

        # Пропускаем заголовок (строка 1), начинаем со строки 2
        for row_idx in range(2, self.source_ws.max_row + 1):

            # Пропускаем скрытые строки (если помечены 'x' в столбце A)
            hide_marker = self.source_ws.cell(row_idx, self.COL_HIDE).value
            if hide_marker and str(hide_marker).strip().lower() == 'x':
                logger.debug(f"⏭️  Пропуск скрытой строки {row_idx}")
                continue

            # Читаем адрес (основной ключ)
            addr = self.source_ws.cell(row_idx, self.COL_ADDR).value

            # Пропускаем пустые строки
            if not addr or str(addr).strip() == '':
                continue

            # Читаем все данные строки
            work_data = {
                'row_num': row_idx,
                'address': str(addr).strip(),
                'start_date': self.source_ws.cell(row_idx, self.COL_START).value,
                'end_date': self.source_ws.cell(row_idx, self.COL_END).value,
                'client': self.source_ws.cell(row_idx, self.COL_CLIENT).value,
                'executor': self.source_ws.cell(row_idx, self.COL_EXEC).value,
                'status': self.source_ws.cell(row_idx, self.COL_STATUS).value,
                'work_name': self.source_ws.cell(row_idx, self.COL_WORK).value,
                'price': self.source_ws.cell(row_idx, self.COL_PRICE).value,
                'report_date': self.source_ws.cell(row_idx, self.COL_DATE).value,
                'transit': self.source_ws.cell(row_idx, self.COL_TRANS).value,
                'note': self.source_ws.cell(row_idx, self.COL_NOTE).value,
                'act_type': self.source_ws.cell(row_idx, self.COL_ACT1).value,
                'act_status': self.source_ws.cell(row_idx, self.COL_ACT2).value,
            }

            # Фильтрация по датам
            if start_date and work_data['start_date']:
                try:
                    work_start = datetime.strptime(str(work_data['start_date']), '%Y-%m-%d')
                    period_start = datetime.strptime(start_date, '%Y-%m-%d')
                    if work_start < period_start:
                        continue
                except:
                    pass

            if end_date and work_data['end_date']:
                try:
                    work_end = datetime.strptime(str(work_data['end_date']), '%Y-%m-%d')
                    period_end = datetime.strptime(end_date, '%Y-%m-%d')
                    if work_end > period_end:
                        continue
                except:
                    pass

            # Фильтрация по статусу
            if filter_status and work_data['status']:
                if str(work_data['status']).strip() != filter_status:
                    continue

            # Фильтрация по исполнителю
            if filter_executor and work_data['executor']:
                if filter_executor not in str(work_data['executor']):
                    continue

            data.append(work_data)

        logger.info(f"✅ Найдено записей: {len(data)}")
        return data

    def generate_report(
        self,
        data: List[Dict],
        output_file: str,
        period_start: str,
        period_end: str,
        client_name: str = "Все клиенты"
    ) -> bool:
        """
        Сгенерировать отчет (ПРИЛОЖЕНИЕ 1, 3, 4)

        Args:
            data: Данные для отчета
            output_file: Путь к выходному файлу
            period_start: Дата начала периода
            period_end: Дата окончания периода
            client_name: Название клиента
        """
        logger.info("=" * 80)
        logger.info("📊 ГЕНЕРАЦИЯ ОТЧЕТА")
        logger.info("=" * 80)
        logger.info(f"📅 Период: {period_start} - {period_end}")
        logger.info(f"👤 Клиент: {client_name}")
        logger.info(f"📝 Записей в отчете: {len(data)}")

        # Создаем новую книгу для отчета
        report_wb = openpyxl.Workbook()
        report_ws = report_wb.active
        report_ws.title = "Отчет"

        # Стили
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Шапка отчета
        current_row = 1

        # Заголовок
        report_ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = report_ws.cell(current_row, 1)
        title_cell.value = "ПРИЛОЖЕНИЕ 1"
        title_cell.font = Font(name='Arial', size=12, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Подзаголовок
        report_ws.merge_cells(f'A{current_row}:E{current_row}')
        subtitle_cell = report_ws.cell(current_row, 1)
        subtitle_cell.value = f"Отчет о выполненных работах"
        subtitle_cell.font = Font(name='Arial', size=11, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # Период
        report_ws.cell(current_row, 1).value = "Период:"
        report_ws.cell(current_row, 2).value = f"{period_start} - {period_end}"
        current_row += 1

        # Клиент
        report_ws.cell(current_row, 1).value = "Клиент:"
        report_ws.cell(current_row, 2).value = client_name
        current_row += 2

        # Заголовки таблицы
        headers = [
            "Адрес + № Задания",
            "Начало работ",
            "Конец работ", 
            "Название работ",
            "Стоимость (руб)"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = report_ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row += 1
        data_start_row = current_row

        # Данные
        total_cost = 0
        for work in data:
            # Основная строка
            report_ws.cell(current_row, 1).value = work['address']
            report_ws.cell(current_row, 2).value = work['start_date']
            report_ws.cell(current_row, 3).value = work['end_date']
            report_ws.cell(current_row, 4).value = work['work_name']
            report_ws.cell(current_row, 5).value = work['price']

            # Стили
            for col_idx in range(1, 6):
                cell = report_ws.cell(current_row, col_idx)
                cell.font = normal_font
                cell.border = border_thin
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Считаем сумму
            if work['price']:
                try:
                    total_cost += float(work['price'])
                except:
                    pass

            current_row += 1

        # Итоговая строка
        current_row += 1
        report_ws.cell(current_row, 1).value = "ИТОГО:"
        report_ws.cell(current_row, 1).font = header_font
        report_ws.cell(current_row, 5).value = total_cost
        report_ws.cell(current_row, 5).font = header_font

        # Ширина столбцов
        report_ws.column_dimensions['A'].width = 60
        report_ws.column_dimensions['B'].width = 15
        report_ws.column_dimensions['C'].width = 15
        report_ws.column_dimensions['D'].width = 30
        report_ws.column_dimensions['E'].width = 15

        # Сохранение
        try:
            report_wb.save(output_file)
            logger.info(f"✅ Отчет сохранен: {output_file}")
            logger.info(f"💰 Общая стоимость: {total_cost} руб.")
            logger.info("=" * 80)
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения: {e}")
            return False


def generate_monthly_report(
    source_file: str,
    output_file: str,
    month: int,
    year: int = 2026,
    client: str = "Все клиенты",
    status_filter: str = None,
    executor_filter: str = None
):
    """
    Сгенерировать месячный отчет

    Args:
        source_file: Путь к рабочей таблице
        output_file: Путь к выходному отчету
        month: Месяц (1-12)
        year: Год
        client: Название клиента для отчета
        status_filter: Фильтр по статусу (опционально)
        executor_filter: Фильтр по исполнителю (опционально)
    """
    from calendar import monthrange

    # Определяем период
    _, last_day = monthrange(year, month)
    period_start = f"{year}-{month:02d}-01"
    period_end = f"{year}-{month:02d}-{last_day}"

    # Создаем генератор
    gen = ReportGeneratorV2(source_file)
    if not gen.load_source():
        return False

    # Извлекаем данные
    data = gen.extract_data(
        start_date=period_start,
        end_date=period_end,
        filter_status=status_filter,
        filter_executor=executor_filter
    )

    if not data:
        logger.warning("⚠️  Нет данных для отчета!")
        return False

    # Генерируем отчет
    return gen.generate_report(
        data=data,
        output_file=output_file,
        period_start=period_start,
        period_end=period_end,
        client_name=client
    )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Генератор отчетов для новой рабочей таблицы")
    parser.add_argument("--source", required=True, help="Путь к рабочей таблице")
    parser.add_argument("--output", required=True, help="Путь к выходному отчету")
    parser.add_argument("--month", type=int, default=1, help="Месяц (1-12)")
    parser.add_argument("--year", type=int, default=2026, help="Год")
    parser.add_argument("--client", default="Все клиенты", help="Название клиента")
    parser.add_argument("--status", default=None, help="Фильтр по статусу")
    parser.add_argument("--executor", default=None, help="Фильтр по исполнителю")

    args = parser.parse_args()

    success = generate_monthly_report(
        source_file=args.source,
        output_file=args.output,
        month=args.month,
        year=args.year,
        client=args.client,
        status_filter=args.status,
        executor_filter=args.executor
    )

    sys.exit(0 if success else 1)
