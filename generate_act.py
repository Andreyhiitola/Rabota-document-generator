#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
generate_act.py - Генерация акта выполненных работ

ОПИСАНИЕ:
- Читает данные из data.xlsx (лист "Работы")
- Берет номер акта из столбца "Номер акта"
- Фильтрует работы по статусу "Выполнен"
- Заполняет шаблон template.xlsx
- Генерирует готовый акт с автоматическим расчётом суммы

ИСПОЛЬЗОВАНИЕ:
    python generate_act.py --data "data.xlsx" --template "template.xlsx"

МАППИНГ ЯЧЕЕК (строка 13+ шаблона):
- ЗАДАНИЕ (левая часть):  A=адрес, B=нач.дата, C=конц.дата, D=вид услуги
- ОТЧЕТ (правая часть):   G=адрес, H=нач.дата, I=конц.дата, J=вид услуги, K=стоимость
"""

import argparse
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings

warnings.filterwarnings("ignore", category=UserWarning)


def sum_to_words(amount: float) -> str:
    """Сумма прописью: 14100 (Четырнадцать тысяч сто) рублей, 00 копеек"""
    rub = int(amount)
    kop = int(round((amount - rub) * 100))

    units = [
        '', 'один', 'два', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь', 'девять',
        'десять', 'одиннадцать', 'двенадцать', 'тринадцать', 'четырнадцать', 'пятнадцать',
        'шестнадцать', 'семнадцать', 'восемнадцать', 'девятнадцать'
    ]
    tens = ['', '', 'двадцать', 'тридцать', 'сорок', 'пятьдесят', 'шестьдесят',
            'семьдесят', 'восемьдесят', 'девяносто']
    hundreds = ['', 'сто', 'двести', 'триста', 'четыреста', 'пятьсот', 'шестьсот',
                'семьсот', 'восемьсот', 'девятьсот']
    thousands = ['', 'тысяча', 'тысячи', 'тысяч']

    def chunk(n: int) -> str:
        if n == 0:
            return ''
        if n < 20:
            return units[n]
        if n < 100:
            return tens[n // 10] + (' ' + units[n % 10] if n % 10 else '')
        return hundreds[n // 100] + (' ' + chunk(n % 100) if n % 100 else '')

    def group(n: int, words: list) -> str:
        if n == 0:
            return ''
        if n == 1:
            return words[1]
        if 2 <= n <= 4:
            return chunk(n) + ' ' + words[2]
        return chunk(n) + ' ' + words[3]

    if rub == 0:
        rub_text = 'ноль'
    else:
        parts = []
        if rub >= 1000:
            t = rub // 1000
            parts.append(group(t, thousands))
            rub %= 1000
        if rub > 0:
            parts.append(chunk(rub))
        rub_text = ' '.join(parts).strip().capitalize()

    return f"{int(amount)} ({rub_text}) рублей, {kop:02d} копеек"


def to_date(val):
    """Excel / pandas дата → dd.mm.yyyy (строка)"""
    if pd.isna(val):
        return ''
    if isinstance(val, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(val))).strftime('%d.%m.%Y')
    if isinstance(val, datetime):
        return val.strftime('%d.%m.%Y')
    return str(val)


def find_sheet(workbook, keywords):
    """Автоматический поиск листа по ключевым словам"""
    sheets = workbook.sheetnames
    for sheet in sheets:
        if any(kw.lower() in sheet.lower() for kw in keywords):
            return sheet
    return sheets[0]


def safe_get(df, row, col, default=''):
    """Безопасное чтение ячейки с дефолтным значением"""
    try:
        val = df.iloc[row, col]
        if pd.isna(val) or str(val).strip() == '':
            return default
        return str(val).strip()
    except Exception:
        return default


def safe_write(sheet, row, col, value):
    """Безопасная запись в ячейку (обход MergedCell)"""
    try:
        cell = sheet.cell(row=row, column=col)
        cell.value = value
        return True
    except Exception:
        return False


def normalize_act_number(value) -> str:
    """Нормализация номера акта: '1' → '1', '1.0' → '1', '01-1' → '1'"""
    s = str(value).strip()
    if s == '' or s.lower() == 'nan':
        return ''

    if '-' in s:
        s = s.split('-')[-1]

    try:
        num = int(float(s))
        return str(num)
    except Exception:
        return s.lstrip('0') or s


def generate_act(data_path: str, template_path: str):
    print("🔄 Загрузка данных...")

    # === 0. Открываем книгу с данными ===
    data_wb = load_workbook(data_path, data_only=True)
    sheets = data_wb.sheetnames
    print(f"📋 Найдено листов: {len(sheets)}")
    
    # === 1. Параметры по умолчанию ===
    month_name = 'январь 2026'      # можно поменять руками
    act_date_str = ''               # пусто = возьмём дату из данных
    status_filter = 'Выполнен'      # фильтр по статусу

    print(f"  Месяц: {month_name}")
    print(f"  Дата акта (фикс.): {act_date_str or '— (будет рассчитана из данных)'}")
    print(f"  Фильтр статуса: {status_filter}")

    # === 2. Лист РАБОТЫ ===
    work_sheet = find_sheet(data_wb, ['работ', 'work', 'основн', 'main'])
    print(f"📊 Данные работ: '{work_sheet}'")

    df_data = pd.read_excel(data_path, sheet_name=work_sheet)
    df_data = df_data.dropna(how='all')
    print(f"  Загружено строк: {len(df_data)}")

    # === 3. БЕРЁМ НОМЕР АКТА ИЗ СТОЛБЦА A (Номер акта) ===
    if 'Номер акта' not in df_data.columns:
        print("❌ ОШИБКА: столбец 'Номер акта' не найден в данных")
        raise SystemExit(1)

    # Берём первое непустое значение из столбца "Номер акта"
    act_numbers = df_data['Номер акта'].dropna()
    if act_numbers.empty:
        print("❌ ОШИБКА: в столбце 'Номер акта' нет значений")
        raise SystemExit(1)

    act_number_raw = str(act_numbers.iloc[0]).strip()
    act_number = normalize_act_number(act_number_raw)

    print(f"  Номер акта (из столбца A): {act_number}")

    # === 4. ФИЛЬТР: статус + номер акта + даты ===
    df_str = df_data.astype(str)
    norm_numbers = df_str['Номер акта'].apply(normalize_act_number)

    mask = (
        df_str['Статус'].str.contains(status_filter, na=False, case=False) &
        (norm_numbers == act_number) &
        pd.notna(df_data['Начало работ']) &
        pd.notna(df_data['Конец работ'])
    )

    df_filtered = df_data[mask].copy()

    if df_filtered.empty:
        print(f"❌ Нет строк для акта {act_number} со статусом '{status_filter}'")
        raise SystemExit(1)

    num_rows = len(df_filtered)
    print(f"  Отфильтровано строк: {num_rows}")

    # === 5. Период по датам ===
    start_date = to_date(df_filtered['Начало работ'].min())
    end_date = to_date(df_filtered['Конец работ'].max())
    print(f"  Период работ: {start_date} - {end_date}")

    # === 6. РАСЦЕНКИ ===
    rates_sheet = find_sheet(data_wb, ['справочник', 'rates', 'расцен', 'расценки'])
    print(f"💰 Расценки: '{rates_sheet}'")
    df_rates = pd.read_excel(data_path, sheet_name=rates_sheet, header=None)

    rates_dict = {}
    for _, row in df_rates.iterrows():
        if pd.notna(row[0]) and pd.notna(row[2]):
            try:
                rates_dict[int(row[0])] = float(row[2])
            except Exception:
                continue
    print(f"  Найдено расценок: {len(rates_dict)}")

    # === 7. РАСЧЁТ СТОИМОСТИ ===
    def get_cost(row):
        if pd.notna(row.get('Стоимость (руб)', pd.NA)):
            try:
                return float(row['Стоимость (руб)'])
            except Exception:
                pass
        desc = str(row.get('Название работ', ''))
        for code, price in rates_dict.items():
            if f"{code}." in desc:
                return price
        return 0.0

    df_filtered['cost'] = df_filtered.apply(get_cost, axis=1)
    total_sum = df_filtered['cost'].sum()
    total_formatted = sum_to_words(total_sum)

    print(f"💵 ИТОГО: {total_sum:,.0f} руб.")
    print(f"  Пропись: {total_formatted}")

    # === 8. ДАТА АКТА ===
    if act_date_str:
        act_date_for_form = act_date_str
    elif 'Дата' in df_filtered.columns:
        act_date_for_form = to_date(df_filtered['Дата'].max())
    else:
        act_date_for_form = end_date
    print(f"  📅 Дата акта: {act_date_for_form}")

    # === 9. ЗАПОЛНЕНИЕ ШАБЛОНА ===
    print("📄 Заполнение шаблона...")
    output_wb = load_workbook(template_path)

    # --- ШАПКА: Номер акта в C5 ---
    for sheet_name in output_wb.sheetnames:
        sheet = output_wb[sheet_name]
        # Пишем номер акта в C5 везде (можно ограничить нужным листом)
        safe_write(sheet, 5, 3, f"№ {act_number}")
        print(f"  Номер акта в C5 листа '{sheet_name}': № {act_number}")

    # --- Лист: Задание_Отчет_Форма_1-3 (ТАБЛИЦА ДАННЫХ) ---
    if 'Задание_Отчет_Форма_1-3' in output_wb.sheetnames:
        sheet13 = output_wb['Задание_Отчет_Форма_1-3']

        FIRST_DATA_ROW = 13

        # === ЗАПОЛНЕНИЕ ДАННЫХ ===
        for idx, (_, row) in enumerate(df_filtered.iterrows()):
            r = FIRST_DATA_ROW + idx

            addr = str(row.get('Адрес + Задание', '')).strip()
            start_work = to_date(row.get('Начало работ', ''))
            end_work = to_date(row.get('Конец работ', ''))
            work_name = str(row.get('Название работ', '')).strip()
            cost = row.get('cost', 0.0)

            # Левая часть (ЗАДАНИЕ): A, B, C, D
            safe_write(sheet13, r, 1, addr)         # A: адрес
            safe_write(sheet13, r, 2, start_work)   # B: дата передачи задания
            safe_write(sheet13, r, 3, end_work)     # C: дата выполнения задания
            safe_write(sheet13, r, 4, work_name)    # D: вид оказанной услуги

            # Правая часть (ОТЧЕТ): G, H, I, J, K
            safe_write(sheet13, r, 7, addr)         # G: адрес
            safe_write(sheet13, r, 8, start_work)   # H: дата передачи задания
            safe_write(sheet13, r, 9, end_work)     # I: дата выполнения задания
            safe_write(sheet13, r, 10, work_name)   # J: вид оказанной услуги
            safe_write(sheet13, r, 11, cost)        # K: стоимость

        print(f"  ✏️ Заполнены {num_rows} строк(и) таблицы")

        # === СУММА (ДИНАМИЧЕСКАЯ) ===
        new_last_data_row = FIRST_DATA_ROW + num_rows - 1
        K_SUM_ROW = new_last_data_row + 1
        J_TEXT_ROW = K_SUM_ROW + 1

        # K_SUM_ROW: динамическая сумма (считает всё выше себя)
        sum_formula = f'=СУММ(K$13:СМЕЩ(K{K_SUM_ROW};-1;0))'
        safe_write(sheet13, K_SUM_ROW, 11, sum_formula)
        print(f"  🔢 K{K_SUM_ROW} = {sum_formula}")

        # J_TEXT_ROW: текст суммы прописью
        safe_write(sheet13, J_TEXT_ROW, 10, total_formatted)
        print(f"  📝 J{J_TEXT_ROW} = {total_formatted}")

        print(f"  ✅ Задание_Отчет_Форма_1-3: готово!")
    else:
        print("  ⚠️ Лист 'Задание_Отчет_Форма_1-3' не найден")

    # === 10. СОХРАНЕНИЕ ===
    now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    out_name = f"Готовый_Акт_{month_name}_{now_str}.xlsx"
    output_wb.save(out_name)

    print("\n🎉 ГОТОВО!")
    print(f"📁 Файл: {out_name}")
    print(f"📋 Номер акта: {act_number}")
    print(f"💰 Итого: {total_formatted}")
    print(f"📈 Работ: {num_rows} шт.")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Генерация акта СМР Северен-Телеком")
    parser.add_argument('--data', required=True, help='Файл данных (*.xlsx)')
    parser.add_argument('--template', required=True, help='Шаблон акта (*.xlsx)')
    args = parser.parse_args()

    try:
        generate_act(args.data, args.template)
    except Exception as e:
        print(f"\n❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        raise SystemExit(1)
