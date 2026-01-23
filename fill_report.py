from flask import Flask, request, jsonify, send_file
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    task_id = data.get('task_number', 'unknown')
    ts = int(datetime.now().timestamp())
    output_name = f"report_{task_id}_{ts}.xlsx"
    output_path = f"/tmp/{output_name}"
    
    template = openpyxl.load_workbook('/app/templates/PRILOZhENIE_1_3_4_ALL_IN_ONE_Zadanie_Otchet_AKT_mesiats_Sentiabr_2025.xlsx')
    ws = template.active
    
    # Точные координаты из твоих скринов
    ws['D15'] = data.get('address', '')
    ws['C16'] = data.get('task_number', '')
    ws['E17'] = data.get('start_date', '')
    ws['K17'] = data.get('end_date', '')
    ws['J18'] = data.get('work_name', '')
    ws['K18'] = data.get('price', '')
    ws['M15'] = data.get('transit_address', '')
    ws['E1'] = datetime.now().strftime('%d.%m.%Y')
    
    template.save(output_path)
    
    return jsonify({
        'status': 'ok',
        'file': output_name,
        'url': f'/download/{output_name}'
    })

@app.route('/download/<filename>')
def download(filename):
    return send_file(f'/tmp/{filename}', as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
