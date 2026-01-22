#!/usr/bin/env python3
"""
Скрипт синхронизации данных из Trello в Excel таблицу
Поддержка: транзитные адреса, автоопределение расценок, метки, статусы

Автор: Система автоматизации "Северен"
Дата: 22.01.2026
"""

import os
import sys
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import logging

try:
    import requests
    import openpyxl
except ImportError as e:
    print(f"❌ Отсутствует модуль: {e}")
    print("   Установите: pip install requests openpyxl")
    sys.exit(1)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/tmp/trello_sync.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ========================================================================
# КОНФИГУРАЦИЯ
# ========================================================================

# Trello API
TRELLO_API_KEY = os.getenv('TRELLO_API_KEY', '')
TRELLO_TOKEN = os.getenv('TRELLO_TOKEN', '')
TRELLO_BOARD_ID = os.getenv('TRELLO_BOARD_ID', 'gfcln6rS')

# Прейскурант (по умолчанию, будет обновлён из Excel)
ПРЕЙСКУРАНТ = {
    "Пункт 1": {
        "название": "Консультации по размещению кабелей",
        "сумма": 1850,
        "ключевые_слова": ["консультация", "согласование доступа", "письмо"]
    },
    "Пункт 2": {
        "название": "Согласование с ЖКС/ГУПРЭП",
        "сумма": 5250,
        "ключевые_слова": ["ЖКС", "ГУПРЭП", "жилкомсервис"]
    },
    "Пункт 3": {
        "название": "Согласование с ТСЖ/УК",
        "сумма": 7050,
        "ключевые_слова": ["ТСЖ", "УК", "управляющая", "замена ВОК", "прокладка кабеля"]
    },
    "Пункт 4": {
        "название": "Транзит/Авария/VIP",
        "сумма": 8600,
        "ключевые_слова": ["транзит", "АВАРИЯ", "срочно", "VIP", "аварийн"]
    },
    "Пункт 5": {
        "название": "Монтаж по фасадам",
        "сумма": 1850,
        "ключевые_слова": ["фасад", "монтаж"]
    },
    "Пункт 6": {
        "название": "Доступ в подвалы/чердаки",
        "сумма": 5250,
        "ключевые_слова": ["подвал", "чердак", "доступ"]
    },
    "Пункт 7": {
        "название": "Доступ в ТЦ/БЦ",
        "сумма": 8600,
        "ключевые_слова": ["ТЦ", "БЦ", "торговый центр", "паркинг", "бизнес центр"]
    }
}

# Списки Trello и их статусы
TRELLO_LISTS_STATUS = {
    "1-й этап. Начало работ.": "В работе",
    "2-й этап. В процессе работы": "В работе",
    "3-й этап СМР": "В работе",
    "4-й этап. Работа сделана": "Выполнено",
    "5-й Этап. Северен остановил работы по своему желанию": "Приостановлено (Северен)",
    "6-й Заказчик Северена остановил работы": "Приостановлено (Заказчик)",
    "7-й ОТКАЗ_РАБОТА ОСТАНОВЛЕНА на 1-м этапе": "Отказ"
}


# ========================================================================
# ПАРСИНГ TRELLO
# ========================================================================

class TrelloParser:
    """Парсер данных из Trello"""
    
    def __init__(self, api_key: str, token: str):
        self.api_key = api_key
        self.token = token
        self.base_url = "https://api.trello.com/1"
        
    def _request(self, endpoint: str, params: Dict = None) -> Dict:
        """Выполнить запрос к API Trello"""
        url = f"{self.base_url}/{endpoint}"
        default_params = {
            'key': self.api_key,
            'token': self.token
        }
        if params:
            default_params.update(params)
            
        try:
            response = requests.get(url, params=default_params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка запроса к Trello API: {e}")
            raise
    
    def get_board_lists(self, board_id: str) -> Dict[str, str]:
        """Получить списки доски (id -> name)"""
        lists = self._request(f"boards/{board_id}/lists")
        return {lst['id']: lst['name'] for lst in lists}
    
    def get_cards(self, board_id: str) -> List[Dict]:
        """Получить все карточки доски"""
        return self._request(
            f"boards/{board_id}/cards",
            {'customFieldItems': 'true', 'attachments': 'true', 'checklists': 'all'}
        )
    
    def parse_card_title(self, title: str) -> Tuple[str, str, List[str]]:
        """
        Парсинг названия карточки
        
        Формат: "[Адрес]. Задание [НОМЕР]. Транзитные адреса: [адрес1], [адрес2]"
        
        Returns:
            (основной_адрес, номер_работы, транзитные_адреса)
        """
        # Извлечение номера работы
        номер_работы = ""
        patterns = [
            r'Номер работы:\s*(\d+)',
            r'Задание\s*№?\s*(\d+)',
            r'Задание:\s*(\d+)',
            r'№\s*(\d+)'
        ]
        for pattern in patterns:
            match = re.search(pattern, title, re.IGNORECASE)
            if match:
                номер_работы = match.group(1)
                break
        
        # Извлечение транзитных адресов
        транзитные_адреса = []
        transit_match = re.search(
            r'Транзитные адреса:\s*(.+?)(?:\.|$)',
            title,
            re.IGNORECASE
        )
        if transit_match:
            transit_str = transit_match.group(1).strip()
            # Разделяем по запятым
            транзитные_адреса = [addr.strip() for addr in transit_str.split(',') if addr.strip()]
        
        # Основной адрес (всё до "Задание" или "Номер работы")
        основной_адрес = title
        for sep in ['Задание', 'Номер работы', 'Транзитные адреса']:
            if sep in title:
                основной_адрес = title.split(sep)[0].strip()
                break
        
        # Удаляем лишние знаки препинания в конце
        основной_адрес = основной_адрес.rstrip('.,;: ')
        
        logger.debug(f"Парсинг названия: '{title}'")
        logger.debug(f"  Адрес: {основной_адрес}")
        logger.debug(f"  Номер: {номер_работы}")
        logger.debug(f"  Транзиты: {транзитные_адреса}")
        
        return основной_адрес, номер_работы, транзитные_адреса
    
    def parse_description(self, description: str) -> Dict[str, str]:
        """
        Парсинг структурированных полей из описания
        
        Ищет поля вида "Название поля: значение"
        """
        result = {}
        
        if not description:
            return result
        
        # Паттерны для извлечения полей
        patterns = {
            'начало_работ': r'Начало работ:\s*(.+)',
            'подрядчик': r'Подрядчик:\s*(.+)',
            'заказчик': r'Заказчик:\s*(.+)',
            'что_делать': r'Что нужно сделать:\s*(.+)',
            'ответственный': r'Ответственный[^:]*:\s*(.+)',
            'договор_ук': r'Нужен ли договор с УК:\s*(.+)',
            'итого': r'ИТОГО:\s*(.+?)(?:руб|$)',
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, description, re.IGNORECASE | re.MULTILINE)
            if match:
                value = match.group(1).strip()
                # Убираем текст после переноса строки
                value = value.split('\n')[0].strip()
                result[key] = value
        
        # Парсинг расценок (если указаны)
        расценки_match = re.search(
            r'Расценки:(.+?)(?=\n\n|\Z)',
            description,
            re.IGNORECASE | re.DOTALL
        )
        if расценки_match:
            расценки_текст = расценки_match.group(1)
            result['расценки_текст'] = расценки_текст
            
            # Извлекаем ИТОГО из расценок
            итого_match = re.search(r'ИТОГО:\s*(\d+)', расценки_текст)
            if итого_match:
                result['итого'] = итого_match.group(1)
        
        # Парсинг дат (конвертация в формат дд.мм.гггг)
        if 'начало_работ' in result:
            result['начало_работ'] = self._normalize_date(result['начало_работ'])
        
        logger.debug(f"Парсинг описания: {result}")
        
        return result
    
    def _normalize_date(self, date_str: str) -> str:
        """Нормализация формата даты"""
        # Попытка парсинга различных форматов
        date_patterns = [
            (r'(\d{2})\.(\d{2})\.(\d{4})', r'\1.\2.\3'),  # дд.мм.гггг
            (r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', r'\1.\2.\3'),  # д.м.гг(гг)
            (r'(\d{4})-(\d{2})-(\d{2})', r'\3.\2.\1'),  # гггг-мм-дд
        ]
        
        for pattern, replacement in date_patterns:
            match = re.search(pattern, date_str)
            if match:
                return re.sub(pattern, replacement, date_str)
        
        return date_str
    
    def parse_checklist_total(self, checklists: List[Dict]) -> Optional[int]:
        """Извлечь ИТОГО из чек-листа"""
        for checklist in checklists:
            for item in checklist.get('checkItems', []):
                name = item.get('name', '')
                # Ищем "ИТОГО: Сумма XXXX рублей"
                match = re.search(r'ИТОГО:\s*(?:Сумма\s*)?(\d+)', name, re.IGNORECASE)
                if match:
                    return int(match.group(1))
        return None
    
    def determine_price(self, card_data: Dict) -> Tuple[str, int]:
        """
        Определить расценку для карточки
        
        Приоритет:
        1. Метка с пунктом прейскуранта
        2. Ключевые слова в описании
        3. По умолчанию Пункт 1
        
        Returns:
            (название_пункта, сумма)
        """
        # Приоритет 1: Метка
        labels = card_data.get('labels', [])
        for label in labels:
            label_name = label.get('name', '')
            for пункт, данные in ПРЕЙСКУРАНТ.items():
                if пункт in label_name:
                    logger.info(f"Расценка из метки: {пункт} = {данные['сумма']} руб.")
                    return пункт, данные['сумма']
        
        # Приоритет 2: Ключевые слова
        description = card_data.get('desc', '').lower()
        название = card_data.get('name', '').lower()
        текст = description + " " + название
        
        for пункт, данные in ПРЕЙСКУРАНТ.items():
            for слово in данные['ключевые_слова']:
                if слово.lower() in текст:
                    logger.info(f"Расценка по ключевому слову '{слово}': {пункт} = {данные['сумма']} руб.")
                    return пункт, данные['сумма']
        
        # Приоритет 3: По умолчанию
        logger.info(f"Расценка по умолчанию: Пункт 1 = 1850 руб.")
        return "Пункт 1", 1850


# ========================================================================
# РАБОТА С EXCEL
# ========================================================================

class ExcelManager:
    """Менеджер для работы с Excel таблицей"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = None
        self.ws_data = None
        self.ws_prices = None
        
    def load(self):
        """Загрузить таблицу"""
        logger.info(f"Загрузка Excel: {self.filepath}")
        try:
            self.wb = openpyxl.load_workbook(self.filepath)
            
            # Поиск листа с данными
            data_sheet_names = ['Данные', 'Data', 'Северен_новая', 'Sheet1']
            for name in self.wb.sheetnames:
                if any(n.lower() in name.lower() for n in data_sheet_names):
                    self.ws_data = self.wb[name]
                    logger.info(f"Найден лист данных: {name}")
                    break
            
            if not self.ws_data:
                self.ws_data = self.wb.active
                logger.warning(f"Используется активный лист: {self.ws_data.title}")
            
            # Поиск листа с прейскурантом
            if 'расценки' in self.wb.sheetnames:
                self.ws_prices = self.wb['расценки']
                logger.info("Найден лист с прейскурантом")
            
        except Exception as e:
            logger.error(f"Ошибка загрузки Excel: {e}")
            raise
    
    def read_prices(self) -> Dict:
        """Прочитать прейскурант из листа 'расценки'"""
        if not self.ws_prices:
            logger.warning("Лист 'расценки' не найден, используется встроенный прейскурант")
            return ПРЕЙСКУРАНТ
        
        прейскурант = {}
        try:
            for row in range(2, self.ws_prices.max_row + 1):
                номер = self.ws_prices.cell(row, 1).value
                описание = self.ws_prices.cell(row, 2).value
                сумма = self.ws_prices.cell(row, 3).value
                
                if номер and сумма:
                    пункт = f"Пункт {номер}"
                    прейскурант[пункт] = {
                        "название": описание or "",
                        "сумма": int(сумма),
                        "ключевые_слова": []
                    }
            
            logger.info(f"Прочитан прейскурант: {len(прейскурант)} позиций")
            return прейскурант
        except Exception as e:
            logger.error(f"Ошибка чтения прейскуранта: {e}")
            return ПРЕЙСКУРАНТ
    
    def find_or_create_row(self, номер_работы: str) -> int:
        """Найти строку с работой или создать новую"""
        # Поиск существующей строки (колонка 6 = адрес + номер работы)
        for row in range(2, self.ws_data.max_row + 1):
            адрес_ячейка = self.ws_data.cell(row, 6).value
            if адрес_ячейка and номер_работы in str(адрес_ячейка):
                logger.info(f"Найдена существующая строка {row} для работы {номер_работы}")
                return row
        
        # Создание новой строки
        new_row = self.ws_data.max_row + 1
        logger.info(f"Создана новая строка {new_row} для работы {номер_работы}")
        return new_row
    
    def write_card_data(self, row: int, card_data: Dict, parsed_data: Dict):
        """Записать данные карточки в Excel"""
        
        # Колонка 1: № (формула)
        self.ws_data.cell(row, 1).value = f"=ROW()-1"
        
        # Колонка 2: Скрыть (x) - оставляем пустым (пользователь заполнит вручную)
        
        # Колонка 3: Акты на закрытие - оставляем пустым
        
        # Колонка 4: Район города (из метки)
        район = parsed_data.get('район', '')
        if район:
            self.ws_data.cell(row, 4).value = район
        
        # Колонка 5: Детали/Архив (полное описание)
        описание = card_data.get('desc', '')
        if описание:
            self.ws_data.cell(row, 5).value = описание
        
        # Колонка 6: Адрес предоставления услуги
        адрес_полный = parsed_data.get('адрес_полный', '')
        if адрес_полный:
            self.ws_data.cell(row, 6).value = адрес_полный
        
        # Колонка 7-8: Период отчета (пользователь заполняет вручную)
        
        # Колонка 9: Стоимость
        стоимость = parsed_data.get('стоимость')
        if стоимость:
            self.ws_data.cell(row, 9).value = int(стоимость)
        
        # Колонка 10: Вид услуги (расценка)
        расценка = parsed_data.get('расценка', '')
        if расценка:
            сумма = parsed_data.get('стоимость', '')
            self.ws_data.cell(row, 10).value = f"{расценка} ({сумма} руб.)"
        
        # Колонка 11: Дата передачи задания (пользователь заполняет)
        
        # Колонка 18: Статус выполнения
        статус = parsed_data.get('статус', '')
        if статус:
            self.ws_data.cell(row, 18).value = статус
        
        # Колонка 21: Инженер Северен (пользователь заполняет)
        
        # Колонка 22: Название клиента
        клиент = parsed_data.get('клиент') or parsed_data.get('заказчик', '')
        if клиент:
            self.ws_data.cell(row, 22).value = клиент
        
        # Колонка 26: Начало работ
        начало_работ = parsed_data.get('начало_работ', '')
        if начало_работ:
            self.ws_data.cell(row, 26).value = начало_работ
        
        # Колонка 27: Дата окончания работ (пользователь заполняет)
        
        # Колонка 28: Исполнитель работ (подрядчик)
        подрядчик = parsed_data.get('подрядчик', '')
        if подрядчик:
            self.ws_data.cell(row, 28).value = подрядчик
        
        logger.info(f"Данные записаны в строку {row}")
    
    def save(self):
        """Сохранить таблицу"""
        try:
            self.wb.save(self.filepath)
            logger.info(f"Таблица сохранена: {self.filepath}")
        except Exception as e:
            logger.error(f"Ошибка сохранения таблицы: {e}")
            raise


# ========================================================================
# ГЛАВНАЯ ФУНКЦИЯ СИНХРОНИЗАЦИИ
# ========================================================================

def sync_trello_to_excel(excel_path: str):
    """Синхронизация данных из Trello в Excel"""
    
    logger.info("=" * 80)
    logger.info("НАЧАЛО СИНХРОНИЗАЦИИ TRELLO → EXCEL")
    logger.info("=" * 80)
    
    # Проверка переменных окружения
    if not TRELLO_API_KEY or not TRELLO_TOKEN:
        logger.error("❌ Не заданы TRELLO_API_KEY или TRELLO_TOKEN")
        logger.error("   Установите переменные окружения!")
        return False
    
    if TRELLO_API_KEY == TRELLO_TOKEN:
        logger.error("❌ TRELLO_API_KEY и TRELLO_TOKEN одинаковые!")
        logger.error("   Проверьте переменные окружения!")
        return False
    
    try:
        # Инициализация
        parser = TrelloParser(TRELLO_API_KEY, TRELLO_TOKEN)
        excel = ExcelManager(excel_path)
        
        # Загрузка Excel
        excel.load()
        
        # Чтение прейскуранта из Excel (если есть)
        global ПРЕЙСКУРАНТ
        прейскурант_из_excel = excel.read_prices()
        if прейскурант_из_excel:
            ПРЕЙСКУРАНТ.update(прейскурант_из_excel)
        
        # Получение списков Trello
        logger.info("Загрузка списков Trello...")
        lists = parser.get_board_lists(TRELLO_BOARD_ID)
        logger.info(f"Найдено списков: {len(lists)}")
        
        # Получение карточек
        logger.info("Загрузка карточек Trello...")
        cards = parser.get_cards(TRELLO_BOARD_ID)
        logger.info(f"Найдено карточек: {len(cards)}")
        
        # Обработка каждой карточки
        обработано = 0
        ошибок = 0
        
        for card in cards:
            try:
                logger.info("-" * 80)
                logger.info(f"Обработка карточки: {card['name']}")
                
                # Парсинг названия
                адрес, номер_работы, транзиты = parser.parse_card_title(card['name'])
                
                if not номер_работы:
                    logger.warning(f"⚠️ Нет номера работы в карточке: {card['name']}")
                    ошибок += 1
                    continue
                
                # Формирование полного адреса
                адрес_полный = адрес
                if номер_работы:
                    адрес_полный += f". Задание {номер_работы}"
                if транзиты:
                    адрес_полный += f". Транзитные адреса: {', '.join(транзиты)}"
                
                # Парсинг описания
                описание_данные = parser.parse_description(card.get('desc', ''))
                
                # Парсинг чек-листа
                итого_чеклист = parser.parse_checklist_total(card.get('checklists', []))
                
                # Определение расценки
                расценка, сумма = parser.determine_price(card)
                
                # Приоритет стоимости: ИТОГО из расценок > ИТОГО из чеклиста > автоопределение
                if 'итого' in описание_данные:
                    стоимость = int(описание_данные['итого'])
                elif итого_чеклист:
                    стоимость = итого_чеклист
                else:
                    стоимость = сумма
                
                # Определение статуса
                list_name = lists.get(card['idList'], '')
                статус = TRELLO_LISTS_STATUS.get(list_name, list_name)
                
                # Извлечение меток
                район = ""
                клиент = ""
                for label in card.get('labels', []):
                    label_name = label.get('name', '')
                    # Район (синие метки или по известным районам)
                    районы = ['Московский', 'Невский', 'Приморский', 'Красногвардейский',
                             'Василеостровский', 'Центральный', 'Колпинский', 'Гатчинский',
                             'Калининский', 'Фрунзенский', 'Петроградский', 'Кировский',
                             'Выборгский', 'Пушкинский']
                    if any(р in label_name for р in районы):
                        район = label_name
                    # Клиент (зелёные метки)
                    клиенты = ['ЭТАЛОН', 'Ростелеком', 'СТОЛОТО', 'Сервис', 'Юнит']
                    if any(к in label_name for к in клиенты):
                        клиент = label_name
                
                # Подготовка данных для записи
                parsed_data = {
                    'адрес': адрес,
                    'номер_работы': номер_работы,
                    'транзиты': транзиты,
                    'адрес_полный': адрес_полный,
                    'район': район,
                    'клиент': клиент,
                    'расценка': расценка,
                    'стоимость': стоимость,
                    'статус': статус,
                    'начало_работ': описание_данные.get('начало_работ', ''),
                    'подрядчик': описание_данные.get('подрядчик', ''),
                    'заказчик': описание_данные.get('заказчик', ''),
                }
                
                # Поиск/создание строки в Excel
                row = excel.find_or_create_row(номер_работы)
                
                # Запись данных
                excel.write_card_data(row, card, parsed_data)
                
                обработано += 1
                logger.info(f"✅ Карточка обработана успешно")
                
            except Exception as e:
                logger.error(f"❌ Ошибка обработки карточки: {e}")
                ошибок += 1
                continue
        
        # Сохранение Excel
        logger.info("-" * 80)
        logger.info("Сохранение изменений...")
        excel.save()
        
        # Итоги
        logger.info("=" * 80)
        logger.info("СИНХРОНИЗАЦИЯ ЗАВЕРШЕНА")
        logger.info(f"✅ Обработано карточек: {обработано}")
        if ошибок > 0:
            logger.warning(f"⚠️ Ошибок: {ошибок}")
        logger.info("=" * 80)
        
        return True
        
    except Exception as e:
        logger.error(f"❌ КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return False


# ========================================================================
# ТОЧКА ВХОДА
# ========================================================================

if __name__ == "__main__":
    # Путь к Excel файлу
    excel_file = os.getenv('EXCEL_FILE_PATH', '/app/excel_files/ПРИЛОЖЕНИЕ_1_3_4_ALL_IN_ONE.xlsx')
    
    if not os.path.exists(excel_file):
        logger.error(f"❌ Файл не найден: {excel_file}")
        sys.exit(1)
    
    success = sync_trello_to_excel(excel_file)
    sys.exit(0 if success else 1)
