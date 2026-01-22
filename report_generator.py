#!/usr/bin/env python3
"""
Генератор отчётов ПРИЛОЖЕНИЕ 1, 3, 4
С автоматической обработкой транзитных адресов и объединением ячеек
"""

import os
import sys
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("❌ Модуль openpyxl не установлен")
    print("   Установите: pip install openpyxl")
    sys.exit(1)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ReportGenerator:
    """Генератор отчётов с обработкой транзитных адресов"""
    
    def __init__(self, source_file: str):
        """
        Args:
            source_file: Путь к файлу с рабочей таблицей
        """
        self.source_file = source_file
        self.wb = None
        self.source_ws = None
        
    def load_source(self):
        """Загрузка исходного файла"""
        logger.info(f"Загрузка файла: {self.source_file}")
        
        if not os.path.exists(self.source_file):
            logger.error(f"❌ Файл не найден: {self.source_file}")
            return False
            
        try:
            self.wb = openpyxl.load_workbook(self.source_file)
            
            # Ищем рабочий лист
            if 'Эксель' in self.wb.sheetnames:
                self.source_ws = self.wb['Эксель']
            elif 'Рабочая таблица' in self.wb.sheetnames:
                self.source_ws = self.wb['Рабочая таблица']
            else:
                self.source_ws = self.wb.active
                
            logger.info(f"✅ Файл загружен. Рабочий лист: {self.source_ws.title}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки файла: {e}")
            return False
    
    def extract_data(self, start_date: str = None, end_date: str = None) -> List[Dict]:
        """
        Извлечение данных из рабочей таблицы
        
        Args:
            start_date: Дата начала периода (формат: ГГГГ-ММ-ДД)
            end_date: Дата конца периода (формат: ГГГГ-ММ-ДД)
            
        Returns:
            Список словарей с данными работ
        """
        logger.info("Извлечение данных из рабочей таблицы...")
        
        data = []
        
        # Определяем колонки (индексы в вашей таблице)
        # Нужно адаптировать под реальную структуру
        COL_ADDR = 5    # E - Адрес
        COL_START = 26  # Z - Начало работ
        COL_CLIENT = 22 # V - Клиент
        COL_PRICE = 9   # I - Стоимость
        COL_SERVICE = 10 # J - Вид услуги
        COL_STATUS = 18  # R - Статус
        COL_NUM = 1      # A - Номер работы
        
        # Проходим по строкам (начиная с 3, т.к. 1-2 заголовки)
        for row_idx in range(3, self.source_ws.max_row + 1):
            
            # Получаем значения ячеек
            addr = self.source_ws.cell(row_idx, COL_ADDR).value
            
            # Пропускаем пустые строки
            if not addr:
                continue
            
            # Проверяем период (если указан)
            # TODO: Добавить фильтрацию по датам
            
            # Извлекаем данные
            work_data = {
                'row_num': row_idx,
                'address': str(addr).strip() if addr else "",
                'work_number': self.source_ws.cell(row_idx, COL_NUM).value,
                'client': self.source_ws.cell(row_idx, COL_CLIENT).value,
                'price': self.source_ws.cell(row_idx, COL_PRICE).value,
                'service': self.source_ws.cell(row_idx, COL_SERVICE).value,
                'status': self.source_ws.cell(row_idx, COL_STATUS).value,
                'start_date': self.source_ws.cell(row_idx, COL_START).value,
                'has_transit': False,
                'transit_addresses': []
            }
            
            # Проверяем есть ли транзитные адреса
            if 'Транзитные адреса:' in work_data['address'] or 'транзит' in work_data['address'].lower():
                work_data['has_transit'] = True
                work_data['transit_addresses'] = self._parse_transit_addresses(work_data['address'])
            
            data.append(work_data)
        
        logger.info(f"✅ Извлечено записей: {len(data)}")
        return data
    
    def _parse_transit_addresses(self, address_text: str) -> List[str]:
        """
        Парсинг транзитных адресов из текста
        
        Args:
            address_text: Текст с адресом и транзитами
            
        Returns:
            Список транзитных адресов
        """
        transit = []
        
        # Ищем "Транзитные адреса:" или "Транзиты:"
        if 'Транзитные адреса:' in address_text:
            parts = address_text.split('Транзитные адреса:')
            if len(parts) > 1:
                transit_text = parts[1].strip()
                # Разделяем по запятым
                transit = [t.strip() for t in transit_text.split(',') if t.strip()]
        
        return transit
    
    def generate_report(
        self,
        data: List[Dict],
        output_file: str,
        period_start: str,
        period_end: str,
        client_name: str = "ПАО Ростелеком"
    ):
        """
        Генерация отчёта ПРИЛОЖЕНИЕ 1, 3, 4
        
        Args:
            data: Список данных работ
            output_file: Путь для сохранения отчёта
            period_start: Дата начала периода
            period_end: Дата окончания периода
            client_name: Название клиента
        """
        logger.info("=" * 80)
        logger.info("ГЕНЕРАЦИЯ ОТЧЁТА")
        logger.info("=" * 80)
        logger.info(f"Период: {period_start} - {period_end}")
        logger.info(f"Клиент: {client_name}")
        logger.info(f"Работ: {len(data)}")
        
        # Создаём новую книгу
        report_wb = openpyxl.Workbook()
        report_ws = report_wb.active
        report_ws.title = "Задание_Отчет"
        
        # Настройка стилей
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        current_row = 1
        report_ws.merge_cells(f'A{current_row}:D{current_row}')
        title_cell = report_ws.cell(current_row, 1)
        title_cell.value = "ПРИЛОЖЕНИЕ № 1 к Договору"
        title_cell.font = Font(name='Arial', size=12, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        report_ws.merge_cells(f'A{current_row}:D{current_row}')
        subtitle_cell = report_ws.cell(current_row, 1)
        subtitle_cell.value = f"ЗАДАНИЕ на оказание услуг"
        subtitle_cell.font = Font(name='Arial', size=11, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Информация о периоде
        report_ws.cell(current_row, 1).value = "Период:"
        report_ws.cell(current_row, 2).value = f"{period_start} - {period_end}"
        current_row += 1
        
        report_ws.cell(current_row, 1).value = "Клиент:"
        report_ws.cell(current_row, 2).value = client_name
        current_row += 2
        
        # Таблица с данными
        # Заголовки колонок
        headers = ["Адрес предоставления услуги", "Дата передачи задания", "Дата выполнения задания", "Вид оказанной услуги"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = report_ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        current_row += 1
        data_start_row = current_row
        
        # Заполняем данные
        for work in data:
            start_row = current_row
            
            # Основной адрес
            report_ws.cell(current_row, 1).value = work['address']
            report_ws.cell(current_row, 2).value = work['start_date']
            report_ws.cell(current_row, 3).value = period_end
            report_ws.cell(current_row, 4).value = work['service']
            
            # Применяем стили
            for col_idx in range(1, 5):
                cell = report_ws.cell(current_row, col_idx)
                cell.font = normal_font
                cell.border = border_thin
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Если есть транзитные адреса - добавляем их
            if work['has_transit'] and work['transit_addresses']:
                rows_to_merge = len(work['transit_addresses']) + 1
                end_row = start_row + rows_to_merge - 1
                
                # Объединяем ячейки для основного адреса
                if rows_to_merge > 1:
                    report_ws.merge_cells(f'A{start_row}:A{end_row}')
                    report_ws.merge_cells(f'B{start_row}:B{end_row}')
                    report_ws.merge_cells(f'C{start_row}:C{end_row}')
                    report_ws.merge_cells(f'D{start_row}:D{end_row}')
                
                # Добавляем транзитные адреса
                current_row += 1
                for transit_addr in work['transit_addresses']:
                    # Транзитные адреса идут отдельными строками
                    # но ячейки уже объединены выше
                    current_row += 1
            else:
                current_row += 1
        
        # Автоширина колонок
        report_ws.column_dimensions['A'].width = 50
        report_ws.column_dimensions['B'].width = 15
        report_ws.column_dimensions['C'].width = 15
        report_ws.column_dimensions['D'].width = 40
        
        # Добавляем выпадающий список для колонки "Вид услуги"
        services_list = [
            "1. Консультации по размещению кабельных линий",
            "2. Согласование с ЖКС/ГУПРЭП",
            "3. Согласование работ по кабельным линиям с ТСЖ/УК",
            "4. Согласование работ по кабельным линиям (транзитные/аварийные)",
            "5. Согласование работ по монтажу по фасадам зданий",
            "6. Согласование доступа в подвалы/чердаки",
            "7. Согласование доступа в ТЦ/БЦ"
        ]
        
        dv = DataValidation(type="list", formula1=f'"{",".join(services_list)}"', allow_blank=True)
        dv.error = 'Выберите значение из списка'
        dv.errorTitle = 'Ошибка ввода'
        report_ws.add_data_validation(dv)
        dv.add(f'D{data_start_row}:D{current_row}')
        
        # Сохранение
        logger.info(f"Сохранение отчёта: {output_file}")
        report_wb.save(output_file)
        logger.info("✅ Отчёт сохранён успешно!")
        
        return True


def generate_monthly_report(
    source_file: str,
    output_file: str,
    month: int,
    year: int = 2026,
    client: str = "ПАО Ростелеком"
):
    """
    Генерация месячного отчёта
    
    Args:
        source_file: Путь к рабочей таблице
        output_file: Путь для сохранения отчёта
        month: Месяц (1-12)
        year: Год
        client: Название клиента
    """
    from calendar import monthrange
    
    # Определяем период
    _, last_day = monthrange(year, month)
    period_start = f"{year}-{month:02d}-01"
    period_end = f"{year}-{month:02d}-{last_day}"
    
    # Создаём генератор
    gen = ReportGenerator(source_file)
    
    if not gen.load_source():
        return False
    
    # Извлекаем данные
    data = gen.extract_data(period_start, period_end)
    
    if not data:
        logger.warning("⚠️ Нет данных для отчёта!")
        return False
    
    # Генерируем отчёт
    return gen.generate_report(
        data=data,
        output_file=output_file,
        period_start=period_start,
        period_end=period_end,
        client_name=client
    )


# ========================================================================
# ТОЧКА ВХОДА
# ========================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Генератор отчётов')
    parser.add_argument('--source', required=True, help='Путь к рабочей таблице')
    parser.add_argument('--output', required=True, help='Путь для сохранения отчёта')
    parser.add_argument('--month', type=int, default=1, help='Месяц (1-12)')
    parser.add_argument('--year', type=int, default=2026, help='Год')
    parser.add_argument('--client', default='ПАО Ростелеком', help='Название клиента')
    
    args = parser.parse_args()
    
    success = generate_monthly_report(
        source_file=args.source,
        output_file=args.output,
        month=args.month,
        year=args.year,
        client=args.client
    )
    
    sys.exit(0 if success else 1)
