#!/usr/bin/env python3
"""
Синхронизация Trello → Excel (НОВАЯ СТРУКТУРА)
14 колонок + справочник работ
"""

import os
import sys
import re
import logging
from datetime import datetime
from typing import Dict, List, Optional

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font
    from dateutil import parser as date_parser
except ImportError as e:
    print(f"❌ Ошибка импорта: {e}")
    print("   Установите: pip install requests openpyxl python-dateutil")
    sys.exit(1)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/tmp/trello_sync.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class TrelloParser:
    """Парсер Trello карточек"""
    
    def __init__(self, api_key: str, token: str, board_id: str):
        self.api_key = api_key
        self.token = token
        self.board_id = board_id
        self.base_url = "https://api.trello.com/1"
        
    def get_cards(self, include_archived: bool = True) -> List[Dict]:
        """
        Получение всех карточек с доски
        
        Args:
            include_archived: Включать ли архивные карточки
        """
        logger.info("Загрузка карточек из Trello...")
        
        # Получаем активные карточки
        url = f"{self.base_url}/boards/{self.board_id}/cards"
        params = {
            'key': self.api_key,
            'token': self.token,
            'fields': 'all',
            'customFieldItems': 'true'
        }
        
        all_cards = []
        
        try:
            # Активные карточки
            response = requests.get(url, params=params)
            response.raise_for_status()
            active_cards = response.json()
            all_cards.extend(active_cards)
            logger.info(f"✅ Загружено активных карточек: {len(active_cards)}")
            
            # Архивные карточки (если нужно)
            if include_archived:
                params['filter'] = 'closed'
                response = requests.get(url, params=params)
                response.raise_for_status()
                archived_cards = response.json()
                all_cards.extend(archived_cards)
                logger.info(f"✅ Загружено архивных карточек: {len(archived_cards)}")
            
            logger.info(f"✅ Всего карточек: {len(all_cards)}")
            return all_cards
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки карточек: {e}")
            return []
    
    def get_labels(self) -> Dict:
        """Получение меток доски"""
        url = f"{self.base_url}/boards/{self.board_id}/labels"
        params = {'key': self.api_key, 'token': self.token}
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            labels_list = response.json()
            
            # Создаём словарь: label_id -> label_name
            labels = {label['id']: label['name'] for label in labels_list}
            return labels
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки меток: {e}")
            return {}
    
    def get_lists(self) -> Dict:
        """Получение списков (колонок) доски"""
        url = f"{self.base_url}/boards/{self.board_id}/lists"
        params = {'key': self.api_key, 'token': self.token}
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            lists_data = response.json()
            
            # Создаём словарь: list_id -> list_name
            lists = {lst['id']: lst['name'] for lst in lists_data}
            return lists
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки списков: {e}")
            return {}
    
    def parse_card(self, card: Dict, labels_map: Dict, lists_map: Dict) -> Dict:
        """
        Парсинг одной карточки
        
        Returns:
            Словарь с данными карточки
        """
        # Название карточки
        name = card.get('name', '')
        
        # Парсим номер работы из названия
        work_number = self._extract_work_number(name)
        
        # Парсим адрес и транзитные адреса
        address, transit_addresses = self._parse_address(name)
        
        # Описание
        description = card.get('desc', '')
        
        # Парсим структурированные поля из описания
        fields = self._parse_description_fields(description)
        
        # Метки
        label_ids = card.get('idLabels', [])
        card_labels = [labels_map.get(lid, '') for lid in label_ids if lid in labels_map]
        
        # Определяем клиента из меток
        client = self._extract_client(card_labels)
        
        # Определяем тип работы из меток или описания
        work_type = self._determine_work_type(card_labels, description, name)
        
        # Статус из списка (колонки)
        list_id = card.get('idList', '')
        status = lists_map.get(list_id, '')
        
        # Проверяем архивирована ли карточка
        is_archived = card.get('closed', False)
        if is_archived:
            status = f"[АРХИВ] {status}" if status else "[АРХИВ]"
        
        # Даты
        start_date = fields.get('Начало работ', '')
        
        # Исполнитель
        executor = fields.get('Подрядчик', fields.get('Исполнитель', ''))
        
        return {
            'work_number': work_number,
            'address': address,
            'transit_addresses': transit_addresses,
            'start_date': start_date,
            'work_type': work_type,
            'client': client,
            'executor': executor,
            'status': status,
            'is_archived': is_archived,
            'description': description,
            'raw_name': name
        }
    
    def _extract_work_number(self, text: str) -> str:
        """Извлечение номера работы"""
        patterns = [
            r'Задание\s*[№#]?\s*(\d+)',
            r'Номер\s+работы[:\s]*(\d+)',
            r'№\s*(\d+)',
            r'\b(\d{5,6})\b'  # 5-6 цифр подряд
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return ""
    
    def _parse_address(self, text: str) -> tuple:
        """
        Парсинг адреса и транзитных адресов
        
        Returns:
            (основной_адрес, список_транзитных_адресов)
        """
        # Ищем транзитные адреса
        transit_pattern = r'Транзитные адреса[:\s]+(.+?)(?:\.|$)'
        transit_match = re.search(transit_pattern, text, re.IGNORECASE | re.DOTALL)
        
        transit_addresses = []
        if transit_match:
            transit_text = transit_match.group(1)
            # Разделяем по запятым
            transit_addresses = [addr.strip() for addr in transit_text.split(',') if addr.strip()]
            
            # Убираем транзитные адреса из основного текста
            main_address = text[:transit_match.start()].strip()
        else:
            main_address = text
        
        # Очищаем адрес от "Задание X", "№X" в конце
        main_address = re.sub(r'\.\s*(Задание|Номер работы|№)[:\s]*\d+.*$', '', main_address, flags=re.IGNORECASE)
        main_address = main_address.strip('. ')
        
        return main_address, transit_addresses
    
    def _parse_description_fields(self, description: str) -> Dict:
        """Парсинг структурированных полей из описания"""
        fields = {}
        
        # Шаблоны для извлечения полей
        patterns = {
            'Начало работ': r'Начало работ[:\s]+([^\n]+)',
            'Подрядчик': r'Подрядчик[:\s]+([^\n]+)',
            'Заказчик': r'Заказчик[:\s]+([^\n]+)',
            'Исполнитель': r'Исполнитель[:\s]+([^\n]+)',
            'Ответственный': r'Ответственный[:\s]+([^\n]+)'
        }
        
        for field_name, pattern in patterns.items():
            match = re.search(pattern, description, re.IGNORECASE)
            if match:
                fields[field_name] = match.group(1).strip()
        
        return fields
    
    def _extract_client(self, labels: List[str]) -> str:
        """Извлечение клиента из меток"""
        # Список известных клиентов
        known_clients = [
            'ЭТАЛОН', 'Ростелеком', 'СТОЛОТО', 
            'Сервис-Недвижимость', 'Юнит Сервис',
            'ПАО "Ростелеком"', 'Сервис Недвижимость'
        ]
        
        for label in labels:
            for client in known_clients:
                if client.lower() in label.lower():
                    return label
        
        return ""
    
    def _determine_work_type(self, labels: List[str], description: str, name: str) -> str:
        """
        Определение типа работы
        
        Returns:
            Название типа работы (из справочника)
        """
        # Сопоставление ключевых слов с типами работ
        work_types_map = {
            '1. Консультации по размещению кабелей ВОЛС': [
                'консультац', 'размещени', 'проклад'
            ],
            '2. Согласование работ по кабельной трассе с ЖКС/ГУПРЭП': [
                'жкс', 'гупрэп', 'жэс'
            ],
            '3. Согласование работ по кабельной трассе с ТСЖ/УК': [
                'тсж', 'управляющ', ' ук '
            ],
            '4. Согласование работ по кабельной трассе (транзитные/аварийные)': [
                'транзит', 'аварий', 'срочн', 'vip'
            ],
            '5. Содействие выполнению монтажных работ по фасадам зданий': [
                'фасад', 'монтаж'
            ],
            '6. Содействие в получении необходимого доступа в подвалы/чердаки': [
                'подвал', 'чердак', 'доступ'
            ],
            '7. Содействие в получении необходимого доступа в ТЦ/БЦ': [
                'тц', 'бц', 'бизнес-центр', 'торговый центр'
            ]
        }
        
        # Проверяем метки
        for label in labels:
            label_lower = label.lower()
            for work_type, keywords in work_types_map.items():
                for keyword in keywords:
                    if keyword in label_lower:
                        return work_type
        
        # Проверяем описание и название
        combined_text = f"{name} {description}".lower()
        for work_type, keywords in work_types_map.items():
            for keyword in keywords:
                if keyword in combined_text:
                    return work_type
        
        # По умолчанию - консультации
        return '1. Консультации по размещению кабелей ВОЛС'


class ExcelManager:
    """Управление Excel файлом"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.ws = None
        
    def load(self) -> bool:
        """Загрузка файла"""
        logger.info(f"Загрузка Excel: {self.file_path}")
        
        if not os.path.exists(self.file_path):
            logger.error(f"❌ Файл не найден: {self.file_path}")
            return False
        
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
            
            # Ищем рабочий лист
            if 'Работы' in self.wb.sheetnames:
                self.ws = self.wb['Работы']
            else:
                self.ws = self.wb.active
            
            logger.info(f"✅ Excel загружен. Лист: {self.ws.title}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки Excel: {e}")
            return False
    
    def find_row_by_work_number(self, work_number: str) -> Optional[int]:
        """
        Найти строку по номеру работы
        
        Args:
            work_number: Номер работы для поиска
            
        Returns:
            Номер строки или None если не найдена
        """
        if not work_number:
            return None
        
        # Ищем в колонке C (Адрес + Задание)
        for row_idx in range(2, self.ws.max_row + 1):
            cell_value = self.ws.cell(row_idx, 3).value  # Колонка C
            if cell_value and work_number in str(cell_value):
                return row_idx
        
        return None
    
    def get_next_empty_row(self) -> int:
        """Получить следующую пустую строку"""
        return self.ws.max_row + 1
    
    def find_or_create_row(self, work_number: str) -> tuple:
        """
        Найти строку по номеру работы или создать новую
        
        Returns:
            (номер_строки, существовала_ли)
        """
        if not work_number:
            # Если нет номера - добавляем в конец
            return self.get_next_empty_row(), False
        
        # Ищем существующую строку
        existing_row = self.find_row_by_work_number(work_number)
        if existing_row:
            return existing_row, True
        
        # Не нашли - создаём новую
        return self.get_next_empty_row(), False
    
    def write_card_data(self, row: int, data: Dict, is_update: bool = False):
        """
        Запись данных карточки в строку
        
        Args:
            row: Номер строки
            data: Данные карточки
            is_update: True если обновляем существующую строку
        """
        # 🔒 ЗАЩИТА: Не обновлять закрытые работы
        date_closed = self.ws.cell(row, 2).value  # Колонка B - Дата закрытия
        if date_closed:
            logger.info(f"  Строка {row}: 🔒 Пропуск (работа закрыта {date_closed})")
            return
        
        action = "Обновление" if is_update else "Создание"
        
        # A: Номер акта - НЕ ТРОГАЕМ при обновлении (заполняется вручную)
        # if not is_update:
        #     self.ws.cell(row, 1).value = ""
        
        # B: Дата закрытия акта - НЕ ТРОГАЕМ (заполняется вручную)
        # self.ws.cell(row, 2).value = None
        
        # C: Адрес + Задание - ВСЕГДА ОБНОВЛЯЕМ
        address_full = data['address']
        if data['work_number']:
            address_full += f". Задание {data['work_number']}"
        self.ws.cell(row, 3).value = address_full
        
        # D: Начало работ - ОБНОВЛЯЕМ если есть
        if data['start_date']:
            try:
                date_obj = date_parser.parse(data['start_date'], dayfirst=True)
                self.ws.cell(row, 4).value = date_obj
            except:
                self.ws.cell(row, 4).value = data['start_date']
        
        # E: Конец работ - НЕ ТРОГАЕМ (заполняется вручную)
        # self.ws.cell(row, 5).value = None
        
        # F: Название работ (тип работы) - ВСЕГДА ОБНОВЛЯЕМ
        self.ws.cell(row, 6).value = data['work_type']
        
        # G: Стоимость - ФОРМУЛА (обновляем только если новая строка)
        if not is_update:
            formula = f'=VLOOKUP(F{row}, Справочник_Работы!$B$3:$C$9, 2, FALSE)'
            self.ws.cell(row, 7).value = formula
        
        # H: Дата формирования отчета - ВСЕГДА обновляем
        self.ws.cell(row, 8).value = datetime.now()
        
        # I: Клиент - ОБНОВЛЯЕМ если есть
        if data['client']:
            self.ws.cell(row, 9).value = data['client']
        
        # J: Исполнитель - ОБНОВЛЯЕМ если есть
        if data['executor']:
            self.ws.cell(row, 10).value = data['executor']
        
        # K: Статус - ВСЕГДА ОБНОВЛЯЕМ (включая [АРХИВ])
        self.ws.cell(row, 11).value = data['status']
        
        # L: Транзитные адреса - ОБНОВЛЯЕМ если есть
        if data['transit_addresses']:
            transit_text = ', '.join(data['transit_addresses'])
            self.ws.cell(row, 12).value = transit_text
        
        # M: Примечание - НЕ ТРОГАЕМ (заполняется вручную)
        # self.ws.cell(row, 13).value = ""
        
        # N: Описание из Trello - ВСЕГДА ОБНОВЛЯЕМ (архив)
        self.ws.cell(row, 14).value = data['description']
        
        # Логирование
        archive_marker = " [АРХИВНАЯ]" if data.get('is_archived', False) else ""
        logger.info(f"  {action} строка {row}: {address_full}{archive_marker}")
    
    def save(self) -> bool:
        """Сохранение файла"""
        try:
            self.wb.save(self.file_path)
            logger.info(f"✅ Файл сохранён: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения: {e}")
            return False


def sync_trello_to_excel(excel_file: str) -> bool:
    """
    Главная функция синхронизации
    
    Args:
        excel_file: Путь к Excel файлу
        
    Returns:
        True если успешно
    """
    logger.info("=" * 80)
    logger.info("СИНХРОНИЗАЦИЯ TRELLO → EXCEL")
    logger.info("=" * 80)
    
    # Получаем переменные окружения
    api_key = os.getenv('TRELLO_API_KEY')
    token = os.getenv('TRELLO_TOKEN')
    board_id = os.getenv('TRELLO_BOARD_ID')
    
    if not all([api_key, token, board_id]):
        logger.error("❌ Не заданы переменные окружения TRELLO_*")
        return False
    
    # Создаём парсер
    parser = TrelloParser(api_key, token, board_id)
    
    # Загружаем данные из Trello
    cards = parser.get_cards()
    if not cards:
        logger.warning("⚠️ Нет карточек для обработки")
        return False
    
    labels_map = parser.get_labels()
    lists_map = parser.get_lists()
    
    # Загружаем Excel
    excel = ExcelManager(excel_file)
    if not excel.load():
        return False
    
    # Обрабатываем карточки
    logger.info(f"Обработка {len(cards)} карточек...")
    
    processed = 0
    updated = 0
    created = 0
    errors = 0
    skipped = 0
    
    for card in cards:
        try:
            # Парсим карточку
            card_data = parser.parse_card(card, labels_map, lists_map)
            
            # Пропускаем если нет номера работы
            if not card_data['work_number']:
                logger.warning(f"⚠️ Пропуск карточки без номера: {card_data['raw_name'][:50]}")
                skipped += 1
                continue
            
            # Находим или создаём строку
            row, exists = excel.find_or_create_row(card_data['work_number'])
            
            # Записываем данные
            excel.write_card_data(row, card_data, is_update=exists)
            
            if exists:
                updated += 1
            else:
                created += 1
            
            processed += 1
            
        except Exception as e:
            logger.error(f"❌ Ошибка обработки карточки: {e}")
            errors += 1
    
    # Сохраняем
    if not excel.save():
        return False
    
    logger.info("=" * 80)
    logger.info("СИНХРОНИЗАЦИЯ ЗАВЕРШЕНА")
    logger.info(f"✅ Всего обработано: {processed}")
    logger.info(f"   - Создано новых: {created}")
    logger.info(f"   - Обновлено: {updated}")
    if skipped > 0:
        logger.warning(f"   - Пропущено (нет номера): {skipped}")
    if errors > 0:
        logger.warning(f"⚠️ Ошибок: {errors}")
    logger.info("=" * 80)
    logger.info("")
    logger.info("💡 ВАЖНО: Строки в Excel НЕ удаляются автоматически!")
    logger.info("   Архивные карточки помечаются как [АРХИВ] в статусе")
    logger.info("=" * 80)
    
    return True


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Синхронизация Trello → Excel')
    parser.add_argument('--file', required=True, help='Путь к Excel файлу')
    
    args = parser.parse_args()
    
    success = sync_trello_to_excel(args.file)
    sys.exit(0 if success else 1)
